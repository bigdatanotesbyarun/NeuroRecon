import google.generativeai as genai
from langchain_google_genai import ChatGoogleGenerativeAI, GoogleGenerativeAIEmbeddings
from langchain.memory import ConversationBufferMemory
from langchain.chains import ConversationalRetrievalChain
from langchain.vectorstores import FAISS
from langchain.text_splitter import CharacterTextSplitter
from rest_framework.decorators import api_view
from rest_framework.response import Response
from langchain.document_loaders import PyPDFLoader
from langchain.schema import Document
import openpyxl 
import os
from pathlib import Path

GENAI_API_KEY = "AIzaSyDLsN1tJQWw00Eh57Rnw_tLLH5VfN5DFkk"
genai.configure(api_key=GENAI_API_KEY)
chat_model = ChatGoogleGenerativeAI(model="gemini-1.5-flash", google_api_key=GENAI_API_KEY)
embeddings = GoogleGenerativeAIEmbeddings(model="models/embedding-001", google_api_key=GENAI_API_KEY)

def read_excel_data(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    excel_data = []
    
    for row in sheet.iter_rows(min_row=2, values_only=True): 
        excel_data.append(row)
        
    return excel_data

def read_excel_data_from_dir(directory_path):
    """
    Reads data from all Excel files in a directory and combines them into a single list.
    
    :param directory_path: Path to the directory containing Excel files
    :return: Combined list of data from all files
    """
    all_data = []
    
    # Check if the directory exists
    if not os.path.exists(directory_path):
        print(f"Directory {directory_path} does not exist.")
        return []
    
    # Iterate over each file in the directory
    for filename in os.listdir(directory_path):
        file_path = os.path.join(directory_path, filename)
        
        # Process only .xlsx files
        if filename.endswith(".xlsx") and os.path.isfile(file_path):
            # Open the Excel file
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            
            # Read data starting from row 2 to skip the header row
            for row in sheet.iter_rows(min_row=2, values_only=True):
                all_data.append(row)  # Add each row to the all_data list
    
    return all_data

def convert_excel_to_documents(excel_data):
    documents = []
    for row in excel_data:
        question = row[0]
        answer = row[1]
        documents.append(f"Q: {question}\nA: {answer}")
    return documents

def create_vector_store(documents):
    text_splitter = CharacterTextSplitter(chunk_size=200, chunk_overlap=20)
    docs = text_splitter.split_text(" ".join(documents))
    doc_objects = [Document(page_content=chunk) for chunk in docs]
    vector_store = FAISS.from_documents(doc_objects, embeddings)
    return vector_store

memory = ConversationBufferMemory(memory_key="chat_history", return_messages=True)

# Define the vector_store inside the view function
@api_view(['POST'])
def get_chat_dataEXCEL(request):
    user_input = request.data.get("message", "")
    if not user_input:
        return Response({"error": "Message is required"}, status=400)

    excel_path = "NeuroReconUI/1.xlsx"  # Provide the correct path to your Excel file
    #excel_data = read_excel_data(excel_path)
    BASE_DIR = Path(__file__).resolve().parent.parent
    path = os.path.join(BASE_DIR, 'Staging', 'DataModel')
    excel_data=read_excel_data_from_dir(path);
    documents = convert_excel_to_documents(excel_data)
    vector_store = create_vector_store(documents)  # Create the vector store inside the request handler

    # Initialize the conversation chain with the newly created vector store
    conversation_chain = ConversationalRetrievalChain.from_llm(
        llm=chat_model,
        retriever=vector_store.as_retriever(),
        memory=memory
    )

    bot_reply = conversation_chain.invoke({
        "question": user_input,
        "chat_history": memory.load_memory_variables({}),
    })

    if isinstance(bot_reply, list) and len(bot_reply) > 0:
        response_text = bot_reply[0].get('answer', 'No Answer available')
    elif isinstance(bot_reply, dict):
        response_text = bot_reply.get('answer', 'No Answer available')
    else:
        response_text = "No Answer available"
    
    memory.save_context({"input": user_input}, {"output": response_text})
    return Response({"response": response_text})
